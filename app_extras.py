from io import BytesIO
from datetime import datetime

from flask import jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

import app_v3

app = app_v3.app
get_db = app_v3.get_db
token_required = app_v3.token_required
roles_required = app_v3.roles_required
row_json = lambda row: app_v3.row_json(row)
from aquagold_validation import coordinates as valid_coordinates, integer as valid_integer


@app.get('/api/reports/insights')
@token_required
def reports_insights():
    with get_db() as db, db.cursor() as cur:
        cur.execute("""
          select c.id,trim(concat_ws(' ',c.first_name,c.last_name)) name,
                 coalesce(sum(v.received_amount),0)::bigint received,count(v.id)::int services
          from customers_v2 c join service_visits v on v.customer_id=c.id
          group by c.id order by received desc limit 10
        """)
        top_customers=[{**row_json(r),'id':str(r['id'])} for r in cur.fetchall()]

        cur.execute("""
          select extract(isodow from (coalesce(visited_at,created_at) at time zone 'Asia/Tehran'))::int weekday,
                 count(*)::int services,coalesce(sum(received_amount),0)::bigint received
          from service_visits group by 1 order by services desc
        """)
        busy_days=[row_json(r) for r in cur.fetchall()]

        cur.execute("""
          select category,count(*)::int count,coalesce(sum(amount),0)::bigint amount
          from expenses group by category order by amount desc
        """)
        expense_categories=[row_json(r) for r in cur.fetchall()]

        cur.execute("""
          select array_to_string((regexp_split_to_array(trim(address),'\\s+'))[1:2],' ') area,
                 count(*)::int customers
          from customers_v2
          where address is not null and trim(address)<>'' and archived=false
          group by 1 order by customers desc limit 10
        """)
        areas=[row_json(r) for r in cur.fetchall()]

        cur.execute("""
          select coalesce(service_type,'نامشخص') service_type,count(*)::int services,
                 coalesce(avg(received_amount),0)::bigint avg_received,
                 coalesce(sum(received_amount),0)::bigint received
          from service_visits group by service_type order by received desc limit 15
        """)
        service_analysis=[row_json(r) for r in cur.fetchall()]

    return jsonify({
        'top_customers':top_customers,
        'busy_days':busy_days,
        'expense_categories':expense_categories,
        'areas':areas,
        'service_analysis':service_analysis,
    })


@app.get('/api/route/nearest')
@token_required
def route_nearest():
    lat, lng = valid_coordinates(request.args.get('lat'), request.args.get('lng'), required=True)
    limit = valid_integer(request.args.get('limit'), 'تعداد مقصد', minimum=1, maximum=25, default=10)
    with get_db() as db, db.cursor() as cur:
        cur.execute("""
          select c.id,trim(concat_ws(' ',c.first_name,c.last_name)) name,c.map_label,c.address,
                 st_y(c.location::geometry) latitude,st_x(c.location::geometry) longitude,
                 round(st_distance(c.location,st_setsrid(st_makepoint(%s,%s),4326)::geography)::numeric,1) distance_m,
                 (select phone from customer_phones p where p.customer_id=c.id order by is_primary desc,id limit 1) phone,
                 (select next_service_at from service_visits v where v.customer_id=c.id and v.next_service_at is not null order by next_service_at desc limit 1) next_service_at
          from customers_v2 c where c.location is not null and c.archived=false
          order by distance_m limit %s
        """,(lng,lat,limit))
        rows=cur.fetchall()
    return jsonify([{**row_json(r),'id':str(r['id'])} for r in rows])


@app.get('/api/audit')
@roles_required('admin')
def audit_list():
    limit = valid_integer(request.args.get('limit'), 'تعداد رویداد', minimum=1, maximum=500, default=100)
    with get_db() as db, db.cursor() as cur:
        cur.execute("select * from audit_log order by created_at desc limit %s",(limit,))
        rows=cur.fetchall()
    return jsonify([row_json(r) for r in rows])


def _money_cell(cell):
    cell.number_format='#,##0'
    cell.alignment=Alignment(horizontal='right')


def _sheet_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font=Font(bold=True)
        cell.alignment=Alignment(horizontal='right')
    ws.freeze_panes='A2'
    ws.sheet_view.rightToLeft=True


def _excel_value(value):
    if isinstance(value, str) and value.lstrip().startswith(('=', '+', '-', '@')):
        return "'" + value
    return value


def _append_safe(ws, values):
    ws.append([_excel_value(value) for value in values])


@app.get('/api/export.xlsx')
@roles_required('admin')
def export_xlsx():
    wb=Workbook()
    ws=wb.active; ws.title='سرویس‌ها'
    _sheet_header(ws,['تاریخ','مشتری','شماره','نوع سرویس','شرح','فاکتور','دریافتی','سهم شرکت','مانده مشتری','روش پرداخت'])
    with get_db() as db, db.cursor() as cur:
        cur.execute("""
          select coalesce(v.visited_at,v.created_at) visit_date,trim(concat_ws(' ',c.first_name,c.last_name)) name,
                 (select phone from customer_phones p where p.customer_id=c.id order by is_primary desc,id limit 1) phone,
                 v.service_type,v.description,v.invoice_amount,v.received_amount,v.company_share_amount,v.customer_balance,v.payment_method
          from service_visits v join customers_v2 c on c.id=v.customer_id order by visit_date desc
        """)
        for r in cur.fetchall():
            _append_safe(ws,[r['visit_date'],r['name'],r['phone'],r['service_type'],r['description'],r['invoice_amount'],r['received_amount'],r['company_share_amount'],r['customer_balance'],r['payment_method']])
            for idx in (6,7,8,9): _money_cell(ws.cell(ws.max_row,idx))

        wc=wb.create_sheet('مشتریان'); _sheet_header(wc,['نام','شماره‌ها','آدرس','نام روی نقشه','پلاک','واحد','مدل دستگاه','Latitude','Longitude'])
        cur.execute("""
          select c.*,case when c.location is null then null else st_y(c.location::geometry) end latitude,
                 case when c.location is null then null else st_x(c.location::geometry) end longitude,
                 coalesce(array_agg(p.phone order by p.is_primary desc,p.id) filter(where p.phone is not null),'{}') phones
          from customers_v2 c left join customer_phones p on p.customer_id=c.id where c.archived=false group by c.id order by c.created_at desc
        """)
        for r in cur.fetchall():
            _append_safe(wc,[' '.join(x for x in [r['first_name'],r['last_name']] if x),' • '.join(r['phones'] or []),r['address'],r['map_label'],r['plaque'],r['unit_no'],r['device_model'],r['latitude'],r['longitude']])

        we=wb.create_sheet('هزینه‌ها'); _sheet_header(we,['تاریخ','دسته','عنوان','مبلغ','توضیحات'])
        cur.execute("select expense_date,category,title,amount,notes from expenses order by expense_date desc")
        for r in cur.fetchall():
            _append_safe(we,[r['expense_date'],r['category'],r['title'],r['amount'],r['notes']]); _money_cell(we.cell(we.max_row,4))

        wf=wb.create_sheet('تسویه شرکت'); _sheet_header(wf,['تاریخ','مبلغ','از دوره','تا دوره','توضیحات'])
        cur.execute("select settled_at,amount,period_from,period_to,notes from company_settlements order by settled_at desc")
        for r in cur.fetchall():
            _append_safe(wf,[r['settled_at'],r['amount'],r['period_from'],r['period_to'],r['notes']]); _money_cell(wf.cell(wf.max_row,2))

    for sheet in wb.worksheets:
        for col in sheet.columns:
            letter=col[0].column_letter
            sheet.column_dimensions[letter].width=min(max(12,max((len(str(c.value)) if c.value is not None else 0) for c in col)+2),45)

    output=BytesIO(); wb.save(output); output.seek(0)
    filename=f"aquagold-report-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output,as_attachment=True,download_name=filename,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
